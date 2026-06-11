import os
import pandas as pd
import numpy as np
from datetime import datetime

# Import openpyxl styles for formatting Excel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_date_safely(date_val):
    """
    Safely parse dates of varying formats (e.g. YYYY-MM-DD, DD/MM/YYYY, Month DD, YYYY)
    into a datetime object. Returns pd.NaT if parsing fails.
    """
    if pd.isna(date_val):
        return pd.NaT
    
    date_str = str(date_val).strip()
    
    # Common format formats to attempt
    formats = [
        "%Y-%m-%d",          # 2025-06-15
        "%d/%m/%Y",          # 15/06/2025
        "%B %d, %Y",         # June 15, 2025
        "%b %d, %Y",         # Jun 15, 2025
        "%d-%m-%Y",          # 15-06-2025
        "%Y/%m/%d"           # 2025/06/15
    ]
    
    for fmt in formats:
        try:
            return pd.to_datetime(date_str, format=fmt)
        except ValueError:
            continue
            
    # Fallback to pandas mixed parser
    try:
        return pd.to_datetime(date_str, errors='coerce')
    except Exception:
        return pd.NaT

def clean_sales_data(sales_df):
    """
    Cleans raw sales transaction data.
    """
    print("Initial raw sales shape:", sales_df.shape)
    
    # 1. Handle Duplicates
    # Drop exact row duplicates
    sales_df = sales_df.drop_duplicates()
    # Drop duplicates on Transaction_ID (keep first instance of transaction)
    sales_df = sales_df.drop_duplicates(subset=["Transaction_ID"], keep="first")
    print("Shape after removing duplicates:", sales_df.shape)
    
    # 2. Standardize Dates
    sales_df["Parsed_Date"] = sales_df["Date"].apply(parse_date_safely)
    # Drop rows with missing dates
    sales_df = sales_df.dropna(subset=["Parsed_Date"])
    # Format standard date string column and timestamp column
    sales_df["Date"] = sales_df["Parsed_Date"].dt.strftime("%Y-%m-%d")
    print("Shape after date parsing and filtering:", sales_df.shape)
    
    # 3. Standardize Categorical Data
    # Product Category cleanup
    sales_df["Product_Category"] = sales_df["Product_Category"].astype(str).str.strip().str.title()
    # Exclude invalid categories
    valid_categories = ["Electronics", "Furniture", "Office Supplies"]
    sales_df = sales_df[sales_df["Product_Category"].isin(valid_categories)]
    
    # Region cleanup
    region_map = {
        'NORTH': 'North', 'N': 'North', 'north': 'North', 'North': 'North',
        'SOUTH': 'South', 'S': 'South', 'south': 'South', 'South': 'South',
        'EAST': 'East', 'E': 'East', 'east': 'East', 'East': 'East',
        'WEST': 'West', 'W': 'West', 'west': 'West', 'West': 'West'
    }
    sales_df["Region"] = sales_df["Region"].astype(str).str.strip().str.upper().map(region_map)
    sales_df = sales_df.dropna(subset=["Region"])
    print("Shape after category & region standardization:", sales_df.shape)
    
    # Customer name fallback
    sales_df["Customer_Name"] = sales_df["Customer_Name"].fillna("Unknown Customer").astype(str).str.strip()
    
    # 4. Handle Missing/Invalid Numeric Data
    # Drop negative units (invalid transactions/returns mapped incorrectly)
    sales_df = sales_df[(sales_df["Units_Sold"] >= 0) | (sales_df["Units_Sold"].isna())]
    
    # Impute numeric fields
    # Case A: If Revenue is null, compute from price and units
    mask_rev_null = sales_df["Revenue"].isna() & sales_df["Units_Sold"].notna() & sales_df["Unit_Price"].notna()
    sales_df.loc[mask_rev_null, "Revenue"] = sales_df.loc[mask_rev_null, "Units_Sold"] * sales_df.loc[mask_rev_null, "Unit_Price"]
    
    # Case B: If Price is null, compute from revenue and units
    mask_price_null = sales_df["Unit_Price"].isna() & sales_df["Revenue"].notna() & sales_df["Units_Sold"].notna() & (sales_df["Units_Sold"] > 0)
    sales_df.loc[mask_price_null, "Unit_Price"] = sales_df.loc[mask_price_null, "Revenue"] / sales_df.loc[mask_price_null, "Units_Sold"]
    
    # Case C: If Units is null, compute from revenue and price
    mask_units_null = sales_df["Units_Sold"].isna() & sales_df["Revenue"].notna() & sales_df["Unit_Price"].notna() & (sales_df["Unit_Price"] > 0)
    sales_df.loc[mask_units_null, "Units_Sold"] = np.round(sales_df.loc[mask_units_null, "Revenue"] / sales_df.loc[mask_units_null, "Unit_Price"])
    
    # Drop any remaining rows that cannot be resolved
    sales_df = sales_df.dropna(subset=["Units_Sold", "Unit_Price", "Revenue"])
    
    # Recalculate/Enforce Revenue accuracy (Revenue = Units * Price)
    sales_df["Units_Sold"] = sales_df["Units_Sold"].astype(int)
    sales_df["Unit_Price"] = sales_df["Unit_Price"].round(2)
    sales_df["Revenue"] = (sales_df["Units_Sold"] * sales_df["Unit_Price"]).round(2)
    
    # Drop temporary parsing column
    sales_df = sales_df.drop(columns=["Parsed_Date"])
    print("Shape after data cleaning and imputation:", sales_df.shape)
    
    return sales_df.reset_index(drop=True)

def apply_transformations(sales_df):
    """
    Applies financial business transformations to cleaned sales transactions.
    """
    # 1. Product specific cost multiplier to model expenses and profits
    # Electronics: ~72% cost, Furniture: ~60% cost, Office Supplies: ~45% cost
    cost_multipliers = {
        "Electronics": 0.72,
        "Furniture": 0.60,
        "Office Supplies": 0.45
    }
    
    # Apply cost ratio
    sales_df["Unit_Cost"] = sales_df.apply(
        lambda row: round(row["Unit_Price"] * cost_multipliers.get(row["Product_Category"], 0.65), 2),
        axis=1
    )
    
    # Compute Total Cost, Profit, and Profit Margin
    sales_df["Cost"] = (sales_df["Units_Sold"] * sales_df["Unit_Cost"]).round(2)
    sales_df["Profit"] = (sales_df["Revenue"] - sales_df["Cost"]).round(2)
    sales_df["Profit_Margin"] = (sales_df["Profit"] / sales_df["Revenue"]).round(4)
    
    # Re-order columns for clarity
    column_order = [
        "Transaction_ID", "Date", "Customer_Name", "Region", 
        "Product_Category", "Product_Name", "Units_Sold", 
        "Unit_Price", "Revenue", "Unit_Cost", "Cost", "Profit", "Profit_Margin"
    ]
    return sales_df[column_order]

def generate_key_metrics(sales_df, targets_df):
    """
    Aggregates cleaned sales and merges targets to generate performance variance reports.
    """
    # Avoid SettingWithCopyWarning
    sales_df = sales_df.copy()
    targets_df = targets_df.copy()
    
    # Standardize Month in targets to YYYY-MM-01 format
    targets_df["Month"] = pd.to_datetime(targets_df["Month"]).dt.strftime("%Y-%m-01")
    
    # Extract month start date from sales transactions
    sales_df["Month_Start"] = pd.to_datetime(sales_df["Date"]).dt.strftime("%Y-%m-01")
    
    # Group actual performance by Month, Region, Product Category
    sales_grouped = sales_df.groupby(["Month_Start", "Region", "Product_Category"]).agg(
        Units_Sold=pd.NamedAgg(column="Units_Sold", aggfunc="sum"),
        Actual_Revenue=pd.NamedAgg(column="Revenue", aggfunc="sum"),
        Total_Cost=pd.NamedAgg(column="Cost", aggfunc="sum"),
        Actual_Profit=pd.NamedAgg(column="Profit", aggfunc="sum")
    ).reset_index()
    
    # Merge Sales Grouped with Targets
    merged_df = pd.merge(
        targets_df,
        sales_grouped,
        left_on=["Month", "Region", "Product_Category"],
        right_on=["Month_Start", "Region", "Product_Category"],
        how="outer"
    )
    
    # Fill target placeholders for any mismatch
    merged_df["Month"] = merged_df["Month"].fillna(merged_df["Month_Start"])
    merged_df["Target_Revenue"] = merged_df["Target_Revenue"].fillna(0.0)
    merged_df["Actual_Revenue"] = merged_df["Actual_Revenue"].fillna(0.0)
    merged_df["Units_Sold"] = merged_df["Units_Sold"].fillna(0).astype(int)
    merged_df["Total_Cost"] = merged_df["Total_Cost"].fillna(0.0)
    merged_df["Actual_Profit"] = merged_df["Actual_Profit"].fillna(0.0)
    
    # Calculate Variance metrics
    merged_df["Revenue_Variance"] = (merged_df["Actual_Revenue"] - merged_df["Target_Revenue"]).round(2)
    merged_df["Target_Achievement_Rate"] = (merged_df["Actual_Revenue"] / merged_df["Target_Revenue"]).round(4)
    # Handle division by zero
    merged_df["Target_Achievement_Rate"] = merged_df["Target_Achievement_Rate"].replace([np.inf, -np.inf], 0.0).fillna(0.0)
    
    merged_df["Profit_Margin"] = (merged_df["Actual_Profit"] / merged_df["Actual_Revenue"]).round(4)
    merged_df["Profit_Margin"] = merged_df["Profit_Margin"].fillna(0.0)
    
    final_cols = [
        "Month", "Region", "Product_Category", "Units_Sold", 
        "Actual_Revenue", "Target_Revenue", "Revenue_Variance", 
        "Target_Achievement_Rate", "Total_Cost", "Actual_Profit", "Profit_Margin"
    ]
    
    # Sort for professional viewing
    merged_df = merged_df[final_cols].sort_values(by=["Month", "Region", "Product_Category"]).reset_index(drop=True)
    return merged_df

def style_excel_sheet(ws, is_metrics=False):
    """
    Applies executive formatting (Navy Theme) to the openpyxl worksheet.
    """
    # Fonts
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Segoe UI", size=10, color="333333")
    total_font = Font(name="Segoe UI", size=10, bold=True, color="000000")
    
    # Fills
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Classic Navy
    total_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # Ice Blue
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    total_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000') # Double bottom line
    )
    
    # Set gridlines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Format Headers
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Format Data Rows
    for r_idx in range(2, ws.max_row + 1):
        for c_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = body_font
            cell.border = thin_border
            
            col_name = ws.cell(row=1, column=c_idx).value
            
            # Alignments
            if col_name in ["Transaction_ID", "Date", "Month", "Region", "Product_Category"]:
                cell.alignment = center_align
            elif col_name in ["Customer_Name", "Product_Name"]:
                cell.alignment = left_align
            else:
                cell.alignment = right_align
                
            # Number Formats
            if col_name in ["Units_Sold"]:
                cell.number_format = '#,##0'
            elif col_name in ["Unit_Price", "Unit_Cost", "Revenue", "Cost", "Profit", "Actual_Revenue", "Target_Revenue", "Revenue_Variance", "Total_Cost", "Actual_Profit"]:
                cell.number_format = '$#,##0.00'
            elif col_name in ["Profit_Margin", "Target_Achievement_Rate"]:
                cell.number_format = '0.0%'
                
    # Add Total Row
    total_row_idx = ws.max_row + 1
    ws.cell(row=total_row_idx, column=1, value="Total / Average").font = total_font
    ws.cell(row=total_row_idx, column=1).alignment = left_align
    ws.cell(row=total_row_idx, column=1).fill = total_fill
    ws.cell(row=total_row_idx, column=1).border = total_border
    
    for c_idx in range(2, ws.max_column + 1):
        col_name = ws.cell(row=1, column=c_idx).value
        cell = ws.cell(row=total_row_idx, column=c_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border
        
        col_letter = get_column_letter(c_idx)
        start_cell = f"{col_letter}2"
        end_cell = f"{col_letter}{total_row_idx - 1}"
        
        if col_name in ["Units_Sold", "Revenue", "Cost", "Profit", "Actual_Revenue", "Target_Revenue", "Revenue_Variance", "Total_Cost", "Actual_Profit"]:
            cell.value = f"=SUM({start_cell}:{end_cell})"
            if col_name in ["Units_Sold"]:
                cell.number_format = '#,##0'
            else:
                cell.number_format = '$#,##0.00'
        elif col_name in ["Unit_Price", "Unit_Cost"]:
            cell.value = f"=AVERAGE({start_cell}:{end_cell})"
            cell.number_format = '$#,##0.00'
        elif col_name == "Profit_Margin":
            # Overall margin = Sum(Profit) / Sum(Revenue)
            # Find Revenue column and Profit column
            rev_col = None
            prof_col = None
            for find_c in range(1, ws.max_column + 1):
                c_name = ws.cell(row=1, column=find_c).value
                if c_name in ["Revenue", "Actual_Revenue"]:
                    rev_col = get_column_letter(find_c)
                if c_name in ["Profit", "Actual_Profit"]:
                    prof_col = get_column_letter(find_c)
            if rev_col and prof_col:
                cell.value = f"={prof_col}{total_row_idx}/{rev_col}{total_row_idx}"
            else:
                cell.value = f"=AVERAGE({start_cell}:{end_cell})"
            cell.number_format = '0.0%'
        elif col_name == "Target_Achievement_Rate":
            # Overall ach rate = Sum(Actual) / Sum(Target)
            act_col = None
            tar_col = None
            for find_c in range(1, ws.max_column + 1):
                c_name = ws.cell(row=1, column=find_c).value
                if c_name == "Actual_Revenue":
                    act_col = get_column_letter(find_c)
                if c_name == "Target_Revenue":
                    tar_col = get_column_letter(find_c)
            if act_col and tar_col:
                cell.value = f"={act_col}{total_row_idx}/{tar_col}{total_row_idx}"
            else:
                cell.value = f"=AVERAGE({start_cell}:{end_cell})"
            cell.number_format = '0.0%'
        else:
            cell.value = ""
            
    # Adjust Columns width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # handle formula representations
            val = str(cell.value or '')
            if val.startswith('='):
                val = "$999,999.00"  # mock width for formulas
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def execute_etl():
    """
    Main ETL workflow.
    """
    print("Initializing Data Cleaning & ETL Pipeline...")
    
    # Define paths
    raw_sales_path = "data/raw/raw_sales_data.csv"
    raw_targets_path = "data/raw/raw_targets_data.csv"
    
    cleaned_dir = "data/cleaned"
    os.makedirs(cleaned_dir, exist_ok=True)
    
    cleaned_sales_csv = os.path.join(cleaned_dir, "cleaned_sales_data.csv")
    cleaned_metrics_csv = os.path.join(cleaned_dir, "cleaned_metrics_data.csv")
    excel_snapshot_path = os.path.join(cleaned_dir, "cleaned_mis_snapshot.xlsx")
    
    # Check if raw files exist
    if not os.path.exists(raw_sales_path) or not os.path.exists(raw_targets_path):
        raise FileNotFoundError("Raw source CSV files missing. Please run `data_generator.py` first.")
        
    # Load raw data
    sales_df = pd.read_csv(raw_sales_path)
    targets_df = pd.read_csv(raw_targets_path)
    
    # 1. Clean Sales
    cleaned_sales = clean_sales_data(sales_df)
    
    # 2. Transform Sales (calculate margins)
    transformed_sales = apply_transformations(cleaned_sales)
    
    # 3. Aggregate metrics and compare targets
    metrics_summary = generate_key_metrics(transformed_sales, targets_df)
    
    # 4. Save cleaned CSVs (for SQL database loading in Step 3)
    transformed_sales.to_csv(cleaned_sales_csv, index=False)
    metrics_summary.to_csv(cleaned_metrics_csv, index=False)
    print(f"Cleaned tables exported to CSV:\n  - {cleaned_sales_csv}\n  - {cleaned_metrics_csv}")
    
    # 5. Export formatted Excel sheets
    print("Writing formatted Excel snapshot...")
    with pd.ExcelWriter(excel_snapshot_path, engine='openpyxl') as writer:
        transformed_sales.to_excel(writer, sheet_name='Cleaned_Sales', index=False)
        metrics_summary.to_excel(writer, sheet_name='Key_Metrics', index=False)
        
        # Style worksheets
        style_excel_sheet(writer.sheets['Cleaned_Sales'])
        style_excel_sheet(writer.sheets['Key_Metrics'], is_metrics=True)
        
    print(f"Excel snapshot successfully created at '{excel_snapshot_path}' with custom executive themes.")
    print("ETL pipeline executed successfully!")

if __name__ == "__main__":
    execute_etl()
